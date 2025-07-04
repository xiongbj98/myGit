# import csv

# # 示例数据：包含多行文本的单元格用 \n 分隔
# data = [
#     ["Name", "Age", "Description"],
#     ["Alice", 30, "第一行\n第二行\n第三行"],
#     ["Bob", 25, "Hello\nWorld"]
# ]

# # 写入 CSV 文件
# with open('output.csv', 'w', newline='', encoding='utf-8') as file:
#     writer = csv.writer(file)
#     writer.writerows(data)

import openpyxl
import datetime

wb = openpyxl.load_workbook("资产汇总.xlsx")
sheet = wb.active
origin_data = {}
for i, row in enumerate(sheet.iter_rows()):
    i += 1
    for data in row:
        if data.row%2:
            # 日期行
            if data.value:
                date = None
                if "," in data.value:
                    date = datetime.datetime.strptime(data.value, "%A, %B %d, %Y")
                else:
                    date = datetime.datetime.strptime(data.value, "%Y年%m月%d日")
                origin_data[date] = sheet[data.column_letter+str(data.row+1)].value


sorted_keys = sorted(origin_data.keys())
sorted_data = {}
for key in sorted_keys:
    year_month_day = datetime.datetime.strftime(key, "%Y_%m_%d")
    year = year_month_day.split('_')[0]
    month = year_month_day.split('_')[1]
    day = year_month_day.split('_')[2]
    # if sorted_data.get(year) is None:
    sorted_data[year] = {} if sorted_data.get(year) is None else sorted_data[year]
    sorted_data[year][month] = {} if sorted_data[year].get(month) is None else sorted_data[year][month]
    sorted_data[year][month][day] = origin_data[key]

# print(sorted_data)
for year, month_data in sorted_data.items():
    sheet_data = []
    header = ['']
    header.extend([str(m)+"月" for m in range(1, 13)])
    sheet_data.append(header)
    for days in range(1,32):
        day_row = [str(days)+"日"]
        day_row.extend(['' for _ in range(12)])
        sheet_data.append(day_row)
    if year not in wb.sheetnames: 
        wb.create_sheet(year)
        year_sheet = wb[year]
    for month, day_data in month_data.items():
        for day, val in day_data.items():
            # print(type(month), type(day))
            sheet_data[int(day)][int(month)] = val
    #         break
    #     break
    # break
    for data_row in sheet_data:
        year_sheet.append(data_row)
    # print(sheet_data)
wb.save("资产汇总_sorted.xlsx")
wb.close()